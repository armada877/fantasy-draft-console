#!/usr/bin/env python3
"""Extract Elboberto's OFFICIAL master projections (2022-2025) from the .xlsm
templates downloaded from his Dropbox. Unlike the user's live-edited league
copies (formulas cached as #DIV/0!), these master files carry fully computed
values: projected FPTS, VBD (start/bench/avg), projected auction $, and tier —
per position sheet (QB/RB/WR/TE/K/DEF).

Output: draft_sheets/elboberto_projections.json
  {year: [{name, pos, fpts, start_vbd, bench_vbd, avg_vbd, proj_value, tier}]}
This is the authoritative baseline for the projected-vs-realized value work.
"""
import json
import os
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
YEARS = [2022, 2023, 2024, 2025, 2026]
POS_SHEETS = ["QB", "RB", "WR", "TE", "K", "DEF"]

# header text -> output field (matched on the header row of each position sheet)
HEADER_MAP = {
    "player": "name", "pos": "pos_sheet", "fpts": "fpts",
    "startvbd": "start_vbd", "benchvbd": "bench_vbd", "avgvbd": "avg_vbd",
    "$": "proj_value", "tier": "tier",
}


def num(v):
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


def find_header_row(ws):
    for r in range(1, 6):
        vals = [str(ws.cell(row=r, column=c).value or "").strip().lower()
                for c in range(1, 40)]
        if "player" in vals and ("fpts" in vals or "$" in vals):
            return r
    return 2


def extract_year(path, pos_label_default):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    rows = []
    for sheet in POS_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hr = find_header_row(ws)
        colmap = {}
        for c in range(1, 40):
            h = str(ws.cell(row=hr, column=c).value or "").strip().lower()
            if h in HEADER_MAP:
                colmap[HEADER_MAP[h]] = c
        if "name" not in colmap or "proj_value" not in colmap:
            continue
        for r in range(hr + 1, ws.max_row + 1):
            name = ws.cell(row=r, column=colmap["name"]).value
            if not name or str(name).strip() == "":
                continue
            rec = {"name": str(name).strip(), "pos": sheet,
                   "proj_value": num(ws.cell(row=r, column=colmap["proj_value"]).value),
                   "fpts": num(ws.cell(row=r, column=colmap.get("fpts", 0)).value) if "fpts" in colmap else None,
                   "start_vbd": num(ws.cell(row=r, column=colmap.get("start_vbd", 0)).value) if "start_vbd" in colmap else None,
                   "bench_vbd": num(ws.cell(row=r, column=colmap.get("bench_vbd", 0)).value) if "bench_vbd" in colmap else None,
                   "avg_vbd": num(ws.cell(row=r, column=colmap.get("avg_vbd", 0)).value) if "avg_vbd" in colmap else None,
                   "tier": ws.cell(row=r, column=colmap["tier"]).value if "tier" in colmap else None}
            rows.append(rec)
    return rows


def main():
    out = {}
    for y in YEARS:
        path = os.path.join(HERE, f"elboberto_{y}.xlsm")
        rows = extract_year(path, y)
        out[str(y)] = rows
        vals = [r["proj_value"] for r in rows if r["proj_value"] is not None]
        pos_counts = {}
        for r in rows:
            pos_counts[r["pos"]] = pos_counts.get(r["pos"], 0) + 1
        print(f"{y}: {len(rows)} players {pos_counts} | proj$ max {max(vals):.0f} "
              f"sum(+) {sum(v for v in vals if v > 0):.0f}")
    path = os.path.join(HERE, "elboberto_projections.json")
    json.dump(out, open(path, "w"), indent=1)
    print("wrote", path)


if __name__ == "__main__":
    main()
