#!/usr/bin/env python3
"""Build draft_sheets/tool_data.json — the data payload the console runs on.

This is the "glue" of the pipeline: it combines
  1. the universal Elboberto projections workbook  (checked-in .xlsm), and
  2. your league's settings + managers            (from scraping/scrape_league.py,
     or the workbook's own defaults if you haven't scraped yet)
into the single JSON the console template expects.

    projections .xlsm ─┐
                       ├─►  tool_data.json  ─►  (inject step)  ─►  console
    scraped league  ───┘

VALUATION REFLECTS YOUR LEAGUE. Rather than trusting the workbook's pre-computed
values (which were baked for whatever settings the author used), this recomputes
per player from YOUR scraped league:
  FPTS  = raw stat projections × your ESPN scoring (statId → points)
  VBD   = FPTS − replacement-level FPTS, where "replacement" accounts for your
          roster: teams × starters[pos], plus your FLEX slots pooled over RB/WR/TE
  worth = $1 + a share of the discretionary auction pool proportional to VBD
So changing FLEX count, scoring (e.g. −1 vs −2 per INT), teams or budget moves the
numbers, as it should. If the workbook lacks raw stat sheets, it falls back to the
sheet's own computed CheatSheet values.

Run:   python3 draft_sheets/build_tool_data.py
Config: config/league.json  (copy config/league.example.json)

────────────────────────────────────────────────────────────────────────────
tool_data.json SCHEMA (what the console template consumes)
────────────────────────────────────────────────────────────────────────────
{
  "me":       "Your Name",              # which manager is you (must be in managers[])
  "budget":   200,                      # auction $ per team
  "starters": {"QB":1,"RB":2,"WR":2,"TE":1},   # required starters by position
  "flex":     2,                        # RB/WR/TE flex slots
  "bench":    6,                        # bench slots
  "my_mult":  {"QB":1,"RB":1,"WR":1,"TE":1},   # your positional value tilt (1 = neutral)
  "players": [                          # the draftable pool
    {"name":"Josh Allen","pos":"QB","tier":"QB1","worth":32,"vbd":89,"fpts":361}
  ],
  "managers": [                         # every team + its bidding tendencies
    {"name":"Your Name","mult":{"QB":1,"RB":1,"WR":1,"TE":1},"conc":50,"maxbuy":200}
  ],
  "keeper_pool": {                      # prior-season rosters priced by the keeper rule
    "Your Name": [{"name":"Bucky Irving","pos":"RB","cost":12,"acq":"TRADE"}]
  }
}

Player fields: worth = projected auction $, vbd = value over replacement (VORP),
fpts = projected season points, tier = e.g. "RB1" (kept from the sheet).

OPPONENT TENDENCIES (mult / conc / maxbuy) drive the bid model in the template:
mult = per-position aggressiveness, conc >72 = stars-and-scrubs tilt, maxbuy = hard
cap on any single bid. A brand-new league has no auction history to calibrate from,
so everyone starts NEUTRAL. **Known gap:** the modeling/simulation pipeline that
calibrates these from years of ESPN auction history is not yet in this repo. Seam:
if config/tendencies.json exists ({"Manager Name": {"mult": {...}, "conc": N,
"maxbuy": N}}), those values are used per manager — that's where calibrated output
plugs in later.
"""
import json
import os
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
POSITIONS = ["QB", "RB", "WR", "TE"]  # skill positions the console drafts

# ESPN lineup-slot ids -> our positions (skill only; K/DST/IR ignored)
ESPN_SLOT = {"0": "QB", "2": "RB", "4": "WR", "6": "TE", "16": "DST", "17": "K"}
FLEX_POS = ("RB", "WR", "TE")          # what a FLEX slot may hold — never K or DST
KDST_SHEET = {"K": "K", "DST": "DEF"}  # console position -> workbook sheet name
ESPN_FLEX_SLOTS = {"23"}   # RB/WR/TE flex
ESPN_POS = {1: "QB", 2: "RB", 3: "WR", 4: "TE", 5: "K", 16: "DST"}   # defaultPositionId
ESPN_BENCH_SLOT = "20"

# ESPN scoring statId -> the stat keys we read off the raw projection sheets
ESPN_STAT = {3: "passYds", 4: "passTD", 20: "passInt", 24: "rushYds", 25: "rushTD",
             42: "recYds", 43: "recTD", 53: "rec", 72: "fumbleLost"}
# used when we can't read ESPN scoring (matches the Elboberto workbook's own scoring)
DEFAULT_SCORING = {"passYds": 0.04, "passTD": 4, "passInt": -1, "rushYds": 0.1,
                   "rushTD": 6, "recYds": 0.1, "recTD": 6, "rec": 0.5, "fumbleLost": -2}
# (group, header) in the raw position sheets -> stat key
STAT_COLS = {("passing", "yds"): "passYds", ("passing", "tds"): "passTD",
             ("rushing", "yds"): "rushYds", ("rushing", "tds"): "rushTD",
             ("receiving", "rec"): "rec", ("receiving", "yds"): "recYds",
             ("receiving", "tds"): "recTD"}


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def load_config():
    path = os.path.join(ROOT, "config", "league.json")
    if not os.path.exists(path):
        sys.exit("Missing config/league.json. Copy config/league.example.json and edit it.")
    with open(path) as f:
        return json.load(f)


def load_tendencies():
    """Optional calibrated opponent profiles ({name: {mult, conc, maxbuy}}); {} if absent."""
    path = os.path.join(ROOT, "config", "tendencies.json")
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return {}


def load_plan():
    """Optional backtest-supported budget plan ({slotKey: $ ceiling}); None if absent.
    The console seeds its per-slot bid ceilings from it. Not a validated-optimal allocation —
    a disciplined default (see analysis/research/strategy_search.py). Absent => neutral frame."""
    path = os.path.join(ROOT, "config", "plan.json")
    if os.path.exists(path):
        with open(path) as f:
            return {k: v for k, v in json.load(f).items() if not k.startswith("_")}
    return None


# ───────────────────── projections: raw stats → FPTS → VBD → $ ─────────────────────
def _wb(xlsm_path):
    if not os.path.exists(xlsm_path):
        sys.exit(f"Projections workbook not found: {xlsm_path}")
    return openpyxl.load_workbook(xlsm_path, data_only=True, read_only=True)


def read_raw_projections(wb):
    """Per player from the QB/RB/WR/TE stat sheets: {name, pos, tier, stored_fpts, stats}.
    Returns None if the raw sheets aren't present (older/other workbook layout)."""
    out = []
    for pos in POSITIONS:
        if pos not in wb.sheetnames:
            continue
        ws = wb[pos]
        ncol = ws.max_column
        groups, headers, cur = [], [], ""
        for c in range(1, ncol + 1):
            g = str(ws.cell(row=1, column=c).value or "").strip().lower()
            cur = g or cur                       # forward-fill merged group labels
            groups.append(cur)
            headers.append(str(ws.cell(row=2, column=c).value or "").strip().lower())
        by_header = {h: i for i, h in enumerate(headers) if h}
        if "player" not in by_header or "fpts" not in by_header:
            continue
        stat_col = {key: i for i, (g, h) in enumerate(zip(groups, headers))
                    if (g, h) in STAT_COLS for key in [STAT_COLS[(g, h)]]}
        stat_col["passInt"] = by_header.get("ints", -1)
        stat_col["fumbleLost"] = by_header.get("fl", -1)
        for r in range(3, ws.max_row + 1):
            name = ws.cell(row=r, column=by_header["player"] + 1).value
            if name is None or str(name).strip() == "":
                continue
            stats = {}
            for key, ci in stat_col.items():
                stats[key] = (_num(ws.cell(row=r, column=ci + 1).value) or 0.0) if ci >= 0 else 0.0
            tier = ws.cell(row=r, column=by_header["tier"] + 1).value if "tier" in by_header else None
            out.append({
                "name": str(name).strip(),
                "pos": pos,
                "tier": str(tier).strip() if tier else pos,
                "stored_fpts": _num(ws.cell(row=r, column=by_header["fpts"] + 1).value) or 0.0,
                "stats": stats,
            })
    return out or None


def read_kdst(wb, teams):
    """Kickers and defences from the workbook's K / DEF sheets: {name, pos, tier, fpts}.

    Both sheets are a flat ranked table -- Rank, name, Pos, ..., fpts -- rather than the
    two-row stat layout the skill positions use, so they get their own reader. Defence names
    are normalised to ESPN's "Falcons D/ST" form so a defence kept from last season still
    matches the scraped roster. Returns [] when the sheets are absent.
    """
    out = []
    for pos, sheet in KDST_SHEET.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        headers = [str(ws.cell(row=1, column=c).value or "").strip().lower()
                   for c in range(1, ws.max_column + 1)]
        try:
            fcol = headers.index("fpts") + 1
        except ValueError:
            continue
        rank = 0
        for r in range(2, ws.max_row + 1):
            name = ws.cell(row=r, column=2).value        # col 1 is Rank, col 2 the name
            fpts = _num(ws.cell(row=r, column=fcol).value)
            if not name or fpts is None:
                continue
            name = str(name).strip()
            if pos == "DST":
                name = name.split()[-1] + " D/ST"        # "Houston Texans" -> "Texans D/ST"
            rank += 1
            out.append({"name": name, "pos": pos, "fpts": float(fpts),
                        "tier": "%s%d" % (pos, (rank - 1) // max(1, teams) + 1)})
    return out


def read_scoring(settings):
    """ESPN statId->points mapped to our stat keys; None if unavailable."""
    items = (settings.get("scoringSettings") or {}).get("scoringItems") or []
    sc = {}
    for it in items:
        key = ESPN_STAT.get(it.get("statId"))
        if key is not None:
            sc[key] = it.get("points", 0) or 0
    return sc or None


def compute_fpts(stats, scoring):
    return sum(stats.get(k, 0.0) * pts for k, pts in scoring.items())


def compute_values(players, starters, flex, bench, teams, budget):
    """Set vbd and worth on each player from the league's roster/teams/budget.

    Works over whatever positions are actually present, so a league that starts a kicker and
    a defence values them on the same footing as the skill positions. Only RB/WR/TE are ever
    promoted through FLEX. Note K and DST points come from the workbook's own projections
    rather than being re-scored through the league's rules -- ESPN's scoring payload carries
    the kicking and defensive stat items separately, and a season that has not been configured
    yet omits them -- so their VBD is directionally right rather than exact. It barely matters:
    with one slot each their replacement level is so close to the starter that they price at
    about a dollar, which is what they actually sell for.
    """
    poss = sorted({x["pos"] for x in players})
    by = {p: sorted([x for x in players if x["pos"] == p], key=lambda x: -x["fpts"])
          for p in poss}
    taken = {p: min(teams * int(starters.get(p, 0)), len(by[p])) for p in poss}
    # FLEX: pool the best remaining RB/WR/TE and promote them to starters
    pool = sorted([x for p in FLEX_POS if p in by for x in by[p][taken[p]:]],
                  key=lambda x: -x["fpts"])
    for x in pool[:teams * int(flex)]:
        taken[x["pos"]] += 1
    baseline = {}                       # replacement = best NON-starter at the position
    for p in poss:
        rest = by[p][taken[p]:]
        baseline[p] = rest[0]["fpts"] if rest else (by[p][-1]["fpts"] if by[p] else 0.0)
    for x in players:
        x["vbd"] = x["fpts"] - baseline[x["pos"]]
    # auction $: reserve $1 for every roster slot league-wide, split the rest by VBD -- but
    # only across the positions the room actually BIDS on. K and DST carry real VBD in points
    # and none of it in dollars: across 2022 and 2025 every kicker in this league sold for
    # exactly $1 (mean $1.05, max $2) and defences averaged $1.3 with a $3 ceiling. Letting
    # their VBD draw from the pool priced the top kicker at $15, which would have read as a
    # standing bargain against a $1 market and pulled money away from the positions that
    # decide the season. They still consume a $1 roster reservation each, which is the whole
    # of their real budget claim.
    roster_slots = sum(int(v) for v in starters.values()) + int(flex) + int(bench)
    discretionary = max(1.0, teams * budget - teams * roster_slots)
    priced = lambda x: x["vbd"] > 0 and x["pos"] in POSITIONS
    total_vbd = sum(x["vbd"] for x in players if priced(x)) or 1.0
    for x in players:
        share = x["vbd"] / total_vbd * discretionary if priced(x) else 0.0
        x["worth"] = 1.0 + share
    return baseline


def read_cheatsheet(wb):
    """Fallback: the workbook's own pre-computed CheatSheet values (NOT league-adjusted)."""
    if "CheatSheet" not in wb.sheetnames:
        return None
    ws = wb["CheatSheet"]
    grid = [[ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            for r in range(1, ws.max_row + 1)]
    ls = wb["LeagueInfo"] if "LeagueInfo" in wb.sheetnames else None
    base = {}
    if ls:
        for r in range(1, ls.max_row + 1):
            pos, val = str(ls.cell(row=r, column=1).value or "").strip(), _num(ls.cell(row=r, column=2).value)
            if pos in POSITIONS and val is not None:
                base[pos] = val
    players = []
    for r, row in enumerate(grid):
        for c, cell in enumerate(row):
            text = str(cell or "")
            pos = text.split(" - ")[0].strip() if "Positional Scarcity" in text else None
            if pos not in POSITIONS:
                continue
            hdr = grid[r + 1]
            col = {str(hdr[c + k] or "").strip().upper(): c + k for k in range(8) if c + k < len(hdr)}
            if "NAME" not in col or "$" not in col:
                continue
            for rr in range(r + 2, len(grid)):
                name = grid[rr][col["NAME"]]
                if name is None or str(name).strip() == "":
                    break
                # The CheatSheet stacks blocks VERTICALLY as well as side by side — TE sits
                # directly under QB in the same columns, separated by its own header rather
                # than a blank row. Reading to the first empty cell therefore swallowed that
                # header as a player called "NAME" and then the whole TE list as quarterbacks,
                # in every season. Stop at the next header instead.
                flat = str(name).strip()
                if flat.upper() == "NAME" or "Positional Scarcity" in flat:
                    break
                worth = _num(grid[rr][col["$"]])
                vbd = _num(grid[rr][col.get("VBD", -1)]) if "VBD" in col else 0.0
                tier = grid[rr][col["TIER"]] if "TIER" in col else None
                players.append({"name": str(name).strip(), "pos": pos,
                                "tier": str(tier).strip() if tier else pos,
                                "worth": max(0, round(worth)) if worth is not None else 0,
                                "vbd": round(vbd or 0.0), "fpts": round((vbd or 0.0) + base.get(pos, 0))})
    return players or None


def workbook_defaults(wb):
    """Roster/budget fallback from the workbook's LeagueInfo sheet (used before you scrape)."""
    ws = wb["LeagueInfo"]
    grid = [[ws.cell(row=r, column=c).value for c in range(1, 7)] for r in range(1, ws.max_row + 1)]
    starters, flex, teams, budget, bench = {}, 0, 12, 200, 6
    for row in grid:
        label, cnt = str(row[2] or "").strip(), _num(row[3])
        if label in POSITIONS and cnt is not None:
            starters[label] = int(cnt)
        elif label.lower().startswith("flex") and cnt is not None:
            flex += int(cnt)
        key, val = str(row[4] or "").strip().lower(), _num(row[5])
        if key == "teams" and val:
            teams = int(val)
        elif key == "budget" and val:
            budget = int(val)
        elif key == "bench size" and val is not None:
            bench = int(val)
    return {"starters": starters or {"QB": 1, "RB": 2, "WR": 2, "TE": 1},
            "flex": flex, "bench": bench, "teams": teams, "budget": budget, "scoring": None}


# ───────────────────────────── league (ESPN scrape) ─────────────────────────────
def read_scraped_league(season):
    path = os.path.join(ROOT, "scraping", "raw", str(season), "league_full.json")
    if not os.path.exists(path):
        return None
    with open(path) as f:
        data = json.load(f)
    if isinstance(data, list):
        data = data[0] if data else {}
    settings = data.get("settings") or {}
    roster = (settings.get("rosterSettings") or {}).get("lineupSlotCounts") or {}
    draft = settings.get("draftSettings") or {}

    starters = {p: 0 for p in sorted(set(ESPN_SLOT.values()))}
    flex = bench = 0
    for slot, cnt in roster.items():
        cnt = int(cnt or 0)
        if slot in ESPN_SLOT:
            starters[ESPN_SLOT[slot]] += cnt
        elif slot in ESPN_FLEX_SLOTS:
            flex += cnt
        elif slot == ESPN_BENCH_SLOT:
            bench += cnt
    starters = {p: n for p, n in starters.items() if n > 0} or {"QB": 1, "RB": 2, "WR": 2, "TE": 1}

    members = {m.get("id"): m for m in (data.get("members") or [])}
    names = []
    owners = {}          # ESPN member GUID -> the console name for that manager
    for t in (data.get("teams") or []):
        m = members.get(t.get("primaryOwner"))
        if m and (m.get("firstName") or m.get("lastName")):
            name = " ".join(f"{m.get('firstName','')} {m.get('lastName','')}".split())
        elif m and m.get("displayName"):
            name = m["displayName"]
        else:
            name = " ".join((t.get("name") or f"{t.get('location','')} {t.get('nickname','')}").split())
        name = name or f"Team {t.get('id')}"
        names.append(name)
        if t.get("primaryOwner"):
            owners[t["primaryOwner"]] = name

    return {
        "starters": starters,
        "flex": flex or 1,
        "bench": bench or 6,
        "budget": int(draft.get("auctionBudget") or 200),
        "managers": names,
        "owners": owners,
        # ESPN serves a stub for a season the commissioner has not rolled over yet: no draft
        # type, and only the handful of scoring items that carry over. Roster slots are missing
        # from it too, so the caller has to know not to trust the shape.
        "configured": bool(draft.get("type")) and len(
            (settings.get("scoringSettings") or {}).get("scoringItems") or []) >= 20,
        "league_name": settings.get("name"),
        "scoring": read_scoring(settings),
    }


def load_player_status():
    """config/player_status.json: {player name: {kind,label,body,news_days}} or {}.

    Written by scraping/scrape_sleeper_status.py. Absent simply means no risk flags — the
    console must build for a league that has never run it.
    """
    path = os.path.join(ROOT, "config", "player_status.json")
    if not os.path.exists(path):
        return {}, None
    try:
        with open(path, encoding="utf-8") as f:
            d = json.load(f)
        return d.get("players") or {}, d.get("pulled")
    except (OSError, ValueError):
        return {}, None


def load_manager_canon():
    """config/manager_canon.json: owner id -> canonical manager name, or {} if absent."""
    path = os.path.join(ROOT, "config", "manager_canon.json")
    if not os.path.exists(path):
        return {}
    try:
        with open(path, encoding="utf-8") as f:
            return {k: v for k, v in json.load(f).items() if not k.startswith("_")}
    except (OSError, ValueError):
        return {}


def read_keeper_pool(prior_season, owners, bump, waiver_value, canon=None):
    """Prior-season ending rosters, priced by the league's keeper rule.

    The rule is bump + what the player was worth last season, where "worth last season" is
    the auction price for a drafted player and a flat waiver value for a free pickup:

        keeper cost = bump + (draft price if drafted/traded else waiver value)

    Two different ESPN fields carry that, and conflating them is the trap:

      playerPoolEntry.keeperValueFuture  what the player cost last season
      acquisitionType                    DRAFT | TRADE | ADD

    For DRAFT and TRADE, keeperValueFuture IS the auction price -- a trade passes the
    original drafter's value to the new team, which is exactly what ESPN records. Verified
    on the 2025 scrape: all 73 DRAFT entries equal the recorded auction cost exactly. For a
    waiver ADD it is an ESPN-computed number unrelated to any bid ($8 for a player nobody
    paid for), so the house waiver value replaces it before the bump is added.

    Returns {console manager name: [{name, pos, cost, acq}]}. Owners are matched by id
    first, then bridged through config/manager_canon.json, because ESPN's member id format
    is NOT stable across seasons -- this league's 2025 ids are brace-wrapped UUIDs while its
    2026 ids are numeric, so an id-only join silently matches nobody. Empty when the prior
    season was never scraped: the console must still build for a brand-new league.
    """
    path = os.path.join(ROOT, "scraping", "raw", str(prior_season), "league_full.json")
    if not os.path.exists(path) or not owners:
        return {}
    try:
        with open(path) as f:
            data = json.load(f)
    except (OSError, ValueError):
        return {}
    if isinstance(data, list):
        data = data[0] if data else {}
    canon = canon or {}
    # canonical name -> the console name this season, so a prior owner id that no longer
    # appears verbatim can still find its manager
    by_canon = {}
    for guid, nm in owners.items():
        c = canon.get(guid)
        if c:
            by_canon[c] = nm

    def console_name(guid):
        if guid in owners:
            return owners[guid]
        c = canon.get(guid)
        return by_canon.get(c) if c else None

    pool = {}
    for t in (data.get("teams") or []):
        name = console_name(t.get("primaryOwner"))
        if not name:
            continue                      # a manager who has since left the league
        rows = []
        for e in ((t.get("roster") or {}).get("entries") or []):
            pe = e.get("playerPoolEntry") or {}
            pl = pe.get("player") or {}
            prior = pe.get("keeperValueFuture")
            if prior is None or not pl.get("fullName"):
                continue
            acq = e.get("acquisitionType") or "DRAFT"
            base = waiver_value if acq == "ADD" else int(prior)
            cost = base + bump
            rows.append({"name": pl["fullName"],
                         "pos": ESPN_POS.get(pl.get("defaultPositionId"), "?"),
                         "cost": max(0, int(cost)), "acq": acq})
        if rows:
            pool[name] = rows
    return pool


def main():
    cfg = load_config()
    me = cfg.get("me") or "Me"
    season = int(cfg.get("season", 2026))
    my_mult = cfg.get("my_mult") or {p: 1.0 for p in POSITIONS}
    wb = _wb(os.path.join(ROOT, cfg.get("projections_xlsm", "")))

    league = read_scraped_league(season)
    if league:
        manager_names = league["managers"]
        print(f"Using scraped league (season {season}): "
              f"{len(manager_names)} managers, ${league['budget']} budget.")
        if me not in manager_names:
            print(f"  WARNING: config 'me' = {me!r} is not among the scraped managers: {manager_names}")
    else:
        league = workbook_defaults(wb)
        n = league["teams"]
        manager_names = [me] + [f"Opponent {i}" for i in range(2, n + 1)]
        print(f"No scrape found (scraping/raw/{season}/league_full.json). "
              f"Using workbook defaults: {n} teams, ${league['budget']} budget, generic opponents.")
        print("  Run scraping/scrape_league.py to pull your real league + managers.")

    # A season ESPN has not been configured for yet reports a partial roster (no K, no D/ST)
    # and a partial scoring table. Say so, and let config state the truth in the meantime.
    if league.get("configured") is False:
        print(f"  ! The {season} league on ESPN looks unconfigured (no draft type, partial "
              "scoring). Roster slots from this scrape are NOT trustworthy — re-scrape once "
              'the season is set up, or set "roster" in config/league.json.')
    roster_cfg = cfg.get("roster") or {}
    if roster_cfg.get("starters"):
        league["starters"] = {k: int(v) for k, v in roster_cfg["starters"].items() if int(v) > 0}
    if roster_cfg.get("flex") is not None:
        league["flex"] = int(roster_cfg["flex"])
    if roster_cfg.get("bench") is not None:
        league["bench"] = int(roster_cfg["bench"])
    if roster_cfg:
        print(f"  Roster shape from config/league.json: {league['starters']} "
              f"+{league['flex']}FLX +{league['bench']}BN.")

    teams = len(manager_names)
    budget = league["budget"]

    # players + valuation
    raw = read_raw_projections(wb)
    if raw is not None:
        scoring = league.get("scoring") or DEFAULT_SCORING
        src = ("your ESPN scoring" if league.get("scoring")
               else "workbook default scoring (no scrape)")
        for x in raw:
            x["fpts"] = compute_fpts(x["stats"], scoring)
        # kickers and defences, but only if this league actually starts them
        kdst = [x for x in read_kdst(wb, teams)
                if int(league["starters"].get(x["pos"], 0)) > 0]
        raw = raw + kdst
        compute_values(raw, league["starters"], league["flex"], league["bench"], teams, budget)
        # trim the full stat sheets (deep waiver fodder) to a sane draftable pool per
        # position — starters + flex + ~2 bench rounds — AFTER valuation so the
        # replacement baseline still sees the whole pool.
        board_pos = [p for p in POSITIONS] + [p for p in KDST_SHEET
                                              if int(league["starters"].get(p, 0)) > 0]
        # K and DST take no FLEX and are never stashed on a bench, so a couple past the
        # starters is the whole realistic pool for them
        cap = {p: teams * (int(league["starters"].get(p, 0))
                           + (int(league["flex"]) + 2 if p in POSITIONS else 0) + 2)
               for p in board_pos}
        raw = [x for p in board_pos
               for x in sorted([y for y in raw if y["pos"] == p], key=lambda y: -y["fpts"])[:cap[p]]]
        players = [{"name": x["name"], "pos": x["pos"], "tier": x["tier"],
                    "worth": max(0, round(x["worth"])), "vbd": round(x["vbd"]),
                    "fpts": round(x["fpts"])} for x in raw]
        print(f"  Valuation recomputed from raw stats × {src}; "
              f"VBD/$ from roster {league['starters']} +{league['flex']}FLX.")
    else:
        players = read_cheatsheet(wb)
        if not players:
            sys.exit("Could not read raw stat sheets or CheatSheet from the workbook.")
        print("  Raw stat sheets absent — using the workbook's pre-computed CheatSheet "
              "values (NOT adjusted to your league).")

    tendencies = load_tendencies()
    if tendencies:
        print(f"  Applying calibrated tendencies from config/tendencies.json "
              f"({len(tendencies)} managers).")
    managers = []
    for name in manager_names:
        t = tendencies.get(name) or {}
        managers.append({
            "name": name,
            "mult": t.get("mult") or {p: 1.0 for p in POSITIONS},
            "conc": t.get("conc", 50),      # <72 => no stars-and-scrubs tilt
            "maxbuy": t.get("maxbuy", budget),
        })

    # keeper costs, so the console can price a keeper instead of asking you to
    keeper_bump = int(cfg.get("keeper_bump", 5) or 0)
    keeper_waiver = int(cfg.get("keeper_waiver_value", 1) or 0)
    keeper_pool = read_keeper_pool(season - 1, league.get("owners") or {},
                                   keeper_bump, keeper_waiver, load_manager_canon())
    if keeper_pool:
        n_add = sum(1 for rows in keeper_pool.values() for r in rows if r["acq"] == "ADD")
        print(f"  Keeper costs from {season - 1} rosters: {len(keeper_pool)} managers, "
              f"${keeper_bump} + prior value; {n_add} waiver adds valued at "
              f"${keeper_waiver} (so ${keeper_waiver + keeper_bump} to keep).")
    elif os.path.exists(os.path.join(ROOT, "scraping", "raw", str(season - 1),
                                     "league_full.json")):
        # the prior season IS scraped, so an empty pool means the owner join failed --
        # say so rather than shipping a console that just never offers a keeper price
        print(f"  ! {season - 1} is scraped but no keeper costs matched a current manager. "
              "ESPN member ids change format between seasons; add the missing ids to "
              "config/manager_canon.json to bridge them.")

    # nomination order, if the league has announced one. Validated against the scraped
    # managers: a name that matches nobody would silently put the wrong team on the clock.
    draft_order = [str(n) for n in (cfg.get("draft_order") or [])]
    unknown = [n for n in draft_order if n not in manager_names]
    if unknown:
        print(f"  ! draft_order names not among the league's managers, dropped: {unknown}")
        draft_order = [n for n in draft_order if n in manager_names]
    if draft_order:
        missing = [n for n in manager_names if n not in draft_order]
        print(f"  Nomination order: {' -> '.join(draft_order)}"
              + (f"  (! not listed: {missing})" if missing else ""))

    # Keepers a manager has actually declared. Checked against that manager's own prior
    # roster, because an announcement the console cannot find is worse than no announcement:
    # it would silently fall back to the projection and look like it had been honoured.
    announced = {}
    for mgr, player in (cfg.get("announced_keepers") or {}).items():
        if mgr not in manager_names:
            print(f"  ! announced_keepers: no manager named {mgr!r} — ignored")
            continue
        roster = keeper_pool.get(mgr) or []
        hit = next((r for r in roster if r["name"].lower() == str(player).strip().lower()), None)
        if not hit:
            print(f"  ! announced_keepers: {player!r} is not on {mgr}'s prior roster — ignored")
            continue
        announced[mgr] = hit["name"]
    if announced:
        print(f"  Announced keepers ({len(announced)}): "
              + ", ".join(f"{m} -> {p}" for m, p in announced.items()))

    status, status_at = load_player_status()
    if status:
        hit = 0
        for p in players:
            r = status.get(p["name"])
            if r:
                p["risk"] = r
                hit += 1
        print(f"  Availability flags on {hit} players (pulled {status_at}).")

    plan = load_plan()
    if plan:
        print(f"  Budget plan from config/plan.json (bid ceilings): "
              f"RB1 ${plan.get('RB1','?')} / WR1 ${plan.get('WR1','?')} / … (backtest-supported).")

    out = {
        "me": me,
        "league_name": league.get("league_name") or cfg.get("league_name"),
        "season": season,
        "budget": budget,
        "starters": league["starters"],
        "flex": league["flex"],
        "bench": league["bench"],
        "my_mult": {p: float(my_mult.get(p, 1.0)) for p in league["starters"]},
        "plan": plan,          # per-slot bid ceilings (backtest-supported plan); None => neutral frame
        "players": players,
        "managers": managers,
        "keeper_pool": keeper_pool,
        "draft_order": draft_order,
        "announced_keepers": announced,
        "status_pulled": status_at,
        "announced_source": cfg.get("announced_keepers_source") or None,
        "announced_pulled": cfg.get("announced_keepers_pulled") or None,
        # lets the console follow the live Sleeper draft; None simply disables that panel
        "sleeper_league_id": str(cfg.get("sleeper_league_id") or "") or None,
    }
    path = os.path.join(HERE, "tool_data.json")
    with open(path, "w") as f:
        json.dump(out, f, separators=(",", ":"))
    counts = {}
    for p in players:
        counts[p["pos"]] = counts.get(p["pos"], 0) + 1
    top = sorted(players, key=lambda p: -p["worth"])[:5]
    print(f"Wrote {path}")
    print(f"  {len(players)} players {counts} | {len(managers)} managers | "
          f"roster {out['starters']} +{out['flex']}FLX +{out['bench']}BN | ${budget} | me={me!r}")
    print("  priciest:", ", ".join(f"{p['name']} ${p['worth']}({p['pos']})" for p in top))
    print("\nNext: inject into the console template (see README 'Quickstart' step 2).")


if __name__ == "__main__":
    main()
